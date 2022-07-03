# 4つのライブラリインストール
import os, sys, glob
sys.path.append(os.path.join(os.path.dirname(__file__), 'site-packages'))
import openpyxl as px

# レポートファイルを作成
NEW_FILE = "report.xlsx"

# 指定した場所のエクセルのファイルを全て取得する
files = glob.glob("./*.xls*")

# 空のエクセルブックを作成
all_data = []
for f in files:
    # 開いているexcelと前に作成したレポートがあっても取り込まない
    if f.startswith('./~')  or f == NEW_FILE:
        continue
    # excelを開く
    wb=px.load_workbook(f, data_only=True)
    # シートを開く
    ws = wb.worksheets[0]
    # シートを読み込んで全行取得する
    for row in ws.iter_rows(min_row=2):
        # 1列目と3列目が空白文字の時には取得せず、次の行へ進む
        if row[0].value is None or \
            row[2].value is None:
            continue
        values = []
        # 全セルを舐め回してデータを取得する
        for col in row:
            values.append(col.value)
        # 全セルデータを一つの配列に保存する
        all_data.append(values)

# ここからall.xlsxを作る作業
# print(all_data)
wb = px.Workbook()
ws = wb.worksheets[0]
start_row = 1
start_col = 1
# 全セルデータを順番に書き込み
for y, row in enumerate(all_data):
    for x, cell in enumerate(row):
        ws.cell(row=start_row + y,
                    column=start_col + x,
                    value=all_data[y][x])

#名前を付けて保存
wb.save(NEW_FILE)
